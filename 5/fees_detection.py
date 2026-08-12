"""
Détection du template "All Fees" dans un classeur uploadé.

⚠️ CORRECTIF : _get_headers_from_rows levait IndexError sur une feuille
chargée en mode read_only (openpyxl.load_workbook(..., read_only=True),
le mode utilisé en vrai dans send_mass_update_data.py) quand une des
lignes demandées (SECTION_TITLE_ROW / COLUMN_LABEL_ROW) n'existe pas
encore dans la feuille — typiquement un fichier ADL qui n'a que sa
ligne 1 remplie. En écriture normale, ws[row_index] auto-étend
silencieusement ; en lecture seule, ws[row_index] fait en interne
tuple(ws.iter_rows(min_row=row_index, max_row=row_index))[0], et si la
ligne n'existe pas, iter_rows renvoie un générateur vide -> IndexError
sur le [0]. Corrigé en attrapant IndexError et en traitant la ligne
comme "aucun header trouvé" plutôt que de laisser planter la détection.
"""

from typing import Iterable, Set

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants.fees_template import COLUMN_LABEL_ROW, HEADERS, SECTION_TITLE_ROW


def _get_headers_from_rows(ws: Worksheet, row_indices: Iterable[int]) -> Set[str]:
    """Récupère l'ensemble des valeurs non vides sur les lignes données.

    Robuste aux feuilles en lecture seule (read_only=True) où une ligne
    demandée peut ne pas exister encore — dans ce cas, elle ne contribue
    simplement aucun header, plutôt que de faire planter la détection.
    """
    headers: Set[str] = set()
    for row_index in row_indices:
        try:
            row_cells = ws[row_index]
        except IndexError:
            # La ligne n'existe pas dans cette feuille (fréquent en mode
            # read_only pour un fichier qui n'a pas ce nombre de lignes
            # d'en-tête, ex: ADL qui n'a que sa ligne 1) — rien à ajouter.
            continue

        headers |= {
            str(cell.value).strip()
            for cell in row_cells
            if cell.value is not None and str(cell.value).strip()
        }
    return headers


def _is_all_fees_worksheet(ws: Worksheet) -> bool:
    """Returns True if the worksheet matches the 'All Fees' mass update
    template structure (tous les HEADERS attendus doivent apparaître
    quelque part dans les lignes 1 et 3 de la feuille)."""
    expected_headers = set(HEADERS)
    found_headers = _get_headers_from_rows(ws, (SECTION_TITLE_ROW, COLUMN_LABEL_ROW))
    return expected_headers.issubset(found_headers)


def is_all_fees_mass_update(wb: Workbook) -> bool:
    """Parcourt les worksheets du classeur, comme get_mu_version, et
    retourne True si l'une d'elles matche la structure 'All Fees'."""
    for ws in wb.worksheets:
        if _is_all_fees_worksheet(ws):
            return True
    return False

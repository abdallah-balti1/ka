import unittest
from io import BytesIO

import openpyxl

from constants.fees_template import HEADERS, N_BASE_COLUMNS, SECTION_GROUPS
from util.mass_update.fees_detection import is_all_fees_mass_update


class TestFeesDetectionReadOnlyRegression(unittest.TestCase):
    """Non-régression : is_all_fees_mass_update levait IndexError sur une
    feuille chargée en read_only=True (le mode réel utilisé dans
    send_mass_update_data.py) quand une des lignes d'en-tête attendues
    (COLUMN_LABEL_ROW=3 notamment) n'existe pas dans le fichier — cas
    typique d'un fichier ADL qui n'a que sa ligne 1 remplie.

    ws[row_index] fait tuple(ws.iter_rows(min_row=row_index,
    max_row=row_index))[0] en interne : en écriture normale ça
    auto-étend silencieusement, en lecture seule iter_rows renvoie un
    générateur vide si la ligne n'existe pas -> IndexError sur le [0].
    """

    def _save_and_reload_read_only(self, wb) -> openpyxl.Workbook:
        buffer = BytesIO()
        wb.save(buffer)
        wb.close()
        buffer.seek(0)
        return openpyxl.load_workbook(buffer, read_only=True, data_only=True)

    def test_adl_like_single_row_workbook_read_only_does_not_raise(self):
        """LE cas qui plantait : un classeur avec une seule ligne d'en-tête
        (comme l'ADL), chargé en read_only, ne doit JAMAIS lever d'exception
        — il doit simplement retourner False."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fees Review"
        adl_headers = [
            "Prod Cd", "Share Cd", "Umbrella name", "Sub-fund name",
            "Max ADL In Fee", "ADL In Fee", "Max ADL Out Fee", "ADL Out Fee",
        ]
        for col_idx, h in enumerate(adl_headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        wb_ro = self._save_and_reload_read_only(wb)

        try:
            result = is_all_fees_mass_update(wb_ro)
        except IndexError:
            self.fail(
                "is_all_fees_mass_update a levé IndexError au lieu de "
                "retourner False sur un fichier à une seule ligne d'en-tête"
            )
        self.assertFalse(result)
        wb_ro.close()

    def test_completely_empty_worksheet_read_only_does_not_raise(self):
        """Feuille totalement vide (aucune cellule jamais écrite),
        chargée en read_only — ne doit pas planter non plus."""
        wb = openpyxl.Workbook()
        wb.active.title = "Fees Review"

        wb_ro = self._save_and_reload_read_only(wb)

        try:
            result = is_all_fees_mass_update(wb_ro)
        except IndexError:
            self.fail("is_all_fees_mass_update a levé IndexError sur une feuille vide")
        self.assertFalse(result)
        wb_ro.close()

    def test_all_fees_workbook_still_detected_true_in_read_only(self):
        """Non-régression inverse : le vrai template All Fees (toutes ses
        lignes présentes) doit toujours être détecté correctement même
        après ce correctif."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fees Review"

        base_labels = HEADERS[:N_BASE_COLUMNS]
        for col_idx, label in enumerate(base_labels, start=1):
            ws.cell(row=1, column=col_idx, value=label)

        for start, end, title, _color in SECTION_GROUPS:
            ws.cell(row=1, column=start, value=title)
            for c in range(start, end + 1):
                ws.cell(row=3, column=c, value=HEADERS[c - 1])

        wb_ro = self._save_and_reload_read_only(wb)

        self.assertTrue(is_all_fees_mass_update(wb_ro))
        wb_ro.close()

    def test_only_section_title_row_missing_does_not_raise(self):
        """Cas limite : seule la ligne 1 manque (fichier qui n'a que la
        ligne 3 remplie, improbable mais à ne pas faire planter)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fees Review"
        ws.cell(row=3, column=1, value="Rate")

        wb_ro = self._save_and_reload_read_only(wb)

        try:
            result = is_all_fees_mass_update(wb_ro)
        except IndexError:
            self.fail("is_all_fees_mass_update a levé IndexError")
        self.assertFalse(result)
        wb_ro.close()


if __name__ == "__main__":
    unittest.main()
